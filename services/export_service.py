import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _clasificar_fila(co, mq135):
    def nivel_co(v):
        if v is None: return -1
        if v < 9: return 0
        if v < 35: return 1
        if v < 60: return 2
        return 3

    def nivel_mq135(v):
        if v is None: return -1
        if v < 800: return 0
        if v < 1200: return 1
        if v < 1500: return 2
        return 3

    peor = max(nivel_co(co), nivel_mq135(mq135))
    etiquetas = {-1: "Sin datos", 0: "Buena", 1: "Moderada", 2: "Mala", 3: "Crítica"}
    colores = {-1: "E5E7EB", 0: "BBF7D0", 1: "FDE68A", 2: "FECACA", 3: "991B1B"}
    return etiquetas[peor], colores[peor]


def exportar_historial_excel(lecturas, device_id):
    """
    Recibe una lista de objetos Lectura y genera un archivo Excel en memoria,
    con encabezado informativo y clasificación de calidad de aire coloreada
    por fila para que sea fácil de leer sin conocimientos técnicos.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Encabezado informativo ---
    ws.cell(row=1, column=1, value=f"Historial de sensores — {device_id}").font = Font(bold=True, size=14, color="1F4E78")
    ws.cell(row=2, column=1, value=f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC").font = Font(size=10, color="64748B")
    ws.cell(row=3, column=1, value=f"Total de lecturas: {len(lecturas)}").font = Font(size=10, color="64748B")

    fila_inicio_tabla = 5
    headers = ["ID", "Device ID", "CO (ppm)", "MQ135", "Calidad del aire", "Latitud", "Longitud", "Fecha"]
    for col_num, texto in enumerate(headers, 1):
        celda = ws.cell(row=fila_inicio_tabla, column=col_num, value=texto)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center")

    fila_actual = fila_inicio_tabla + 1
    for i, lectura in enumerate(lecturas):
        etiqueta, color_hex = _clasificar_fila(lectura.co, lectura.mq135)
        valores = [
            lectura.id,
            lectura.device_id,
            lectura.co if lectura.co is not None else "S/D",
            lectura.mq135 if lectura.mq135 is not None else "S/D",
            etiqueta,
            lectura.lat if lectura.lat is not None else "S/D",
            lectura.lng if lectura.lng is not None else "S/D",
            lectura.timestamp.strftime("%Y-%m-%d %H:%M:%S") if lectura.timestamp else "S/D",
        ]
        for col_num, valor in enumerate(valores, 1):
            celda = ws.cell(row=fila_actual, column=col_num, value=valor)
            # Fila alternada para legibilidad
            if i % 2 == 1:
                celda.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        # La columna "Calidad del aire" siempre se resalta con su color de nivel
        celda_nivel = ws.cell(row=fila_actual, column=5)
        celda_nivel.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        if color_hex == "991B1B":
            celda_nivel.font = Font(color="FFFFFF", bold=True)
        else:
            celda_nivel.font = Font(bold=True)
        fila_actual += 1

    ws.freeze_panes = f"A{fila_inicio_tabla + 1}"

    for columna in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_monitoreo_excel(monitoreo, lecturas):
    """Genera un Excel con dos hojas: Resumen (métricas del monitoreo zonal)
    y Recorrido (cada punto GPS + sensores registrados durante la sesión)."""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Hoja Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    duracion_min = None
    if monitoreo.hora_fin and monitoreo.hora_inicio:
        duracion_min = round((monitoreo.hora_fin - monitoreo.hora_inicio).total_seconds() / 60, 1)

    filas_resumen = [
        ("Robot (device_id)", monitoreo.device_id),
        ("Nombre del monitoreo", monitoreo.nombre or "S/D"),
        ("Estado", monitoreo.estado),
        ("Hora de inicio", monitoreo.hora_inicio.strftime("%Y-%m-%d %H:%M:%S") if monitoreo.hora_inicio else "S/D"),
        ("Hora de fin", monitoreo.hora_fin.strftime("%Y-%m-%d %H:%M:%S") if monitoreo.hora_fin else "S/D"),
        ("Duración (minutos)", duracion_min if duracion_min is not None else "S/D"),
        ("Ubicación de inicio", f"{monitoreo.lat_inicio}, {monitoreo.lng_inicio}" if monitoreo.lat_inicio else "S/D"),
        ("Ubicación de fin", f"{monitoreo.lat_fin}, {monitoreo.lng_fin}" if monitoreo.lat_fin else "S/D"),
        ("Promedio CO (ppm)", monitoreo.promedio_co if monitoreo.promedio_co is not None else "S/D"),
        ("Promedio MQ135", monitoreo.promedio_mq135 if monitoreo.promedio_mq135 is not None else "S/D"),
        ("Nivel de calidad de aire", monitoreo.nivel_color or "S/D"),
        ("Centro de la zona (lat, lng)", f"{monitoreo.centro_lat}, {monitoreo.centro_lng}" if monitoreo.centro_lat else "S/D"),
        ("Radio de la zona (metros)", monitoreo.radio_metros if monitoreo.radio_metros is not None else "S/D"),
        ("Distancia total recorrida (m)", monitoreo.distancia_total_m if monitoreo.distancia_total_m is not None else "S/D"),
        ("Velocidad promedio (km/h)", monitoreo.velocidad_promedio_kmh if monitoreo.velocidad_promedio_kmh is not None else "S/D"),
        ("Total de puntos registrados", monitoreo.total_puntos or 0),
    ]

    ws_resumen.cell(row=1, column=1, value="Reporte de Monitoreo Zonal - WALLY").font = Font(bold=True, size=14, color="1F4E78")
    ws_resumen.merge_cells("A1:B1")

    for i, (etiqueta, valor) in enumerate(filas_resumen, start=3):
        celda_etiqueta = ws_resumen.cell(row=i, column=1, value=etiqueta)
        celda_etiqueta.font = Font(bold=True)
        ws_resumen.cell(row=i, column=2, value=valor)

    ws_resumen.column_dimensions["A"].width = 32
    ws_resumen.column_dimensions["B"].width = 40

    # --- Hoja Recorrido ---
    ws_rec = wb.create_sheet("Recorrido")
    headers = ["#", "Latitud", "Longitud", "GPS interpolado", "CO (ppm)", "MQ135", "Velocidad (km/h)", "Fecha/Hora"]
    ws_rec.append(headers)
    for col_num, _ in enumerate(headers, 1):
        celda = ws_rec.cell(row=1, column=col_num)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center")

    for idx, l in enumerate(lecturas, start=1):
        ws_rec.append([
            idx,
            l.lat if l.lat is not None else "S/D",
            l.lng if l.lng is not None else "S/D",
            "Sí" if l.gps_interpolado else "No",
            l.co if l.co is not None else "S/D",
            l.mq135 if l.mq135 is not None else "S/D",
            l.velocidad_kmh if l.velocidad_kmh is not None else "S/D",
            l.timestamp.strftime("%Y-%m-%d %H:%M:%S") if l.timestamp else "S/D",
        ])

    for columna in ws_rec.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in columna)
        ws_rec.column_dimensions[columna[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_usuarios_excel(usuarios):
    """
    Recibe una lista de objetos Usuario y genera un Excel con el listado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["ID", "Nombre", "Email", "Rol", "Verificado", "Creado en"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center")

    for u in usuarios:
        ws.append([
            u.id,
            u.nombre,
            u.email,
            u.rol,
            "Sí" if u.email_verificado else "No",
            u.creado_en.strftime("%Y-%m-%d %H:%M:%S") if u.creado_en else "S/D",
        ])

    for columna in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer